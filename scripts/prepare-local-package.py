"""Build a read-only, local-only preview from a user supplied response ZIP.

Original Office files are copied byte-for-byte. No macros, external relationships,
formulas or archive instructions are executed. This does not create backend versions.
"""
import argparse
import base64
import hashlib
import html
import io
import json
import posixpath
from pathlib import Path
import zipfile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

GROUPS = {
    "商务文件": "business",
    "技术文件": "technical",
    "价格文件": "price",
    "内部管理文件": "internal",
}
W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
R = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
A = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
V = "{urn:schemas-microsoft-com:vml}"

STYLE = """
*{box-sizing:border-box}body{margin:0;background:#edf1f4;color:#1d2939;font:15px/1.8 'Microsoft YaHei',sans-serif}
main{max-width:1080px;margin:20px auto;padding:38px 44px;background:white;box-shadow:0 2px 12px #152c3912;overflow-wrap:anywhere}
h1{font-size:22px;margin:0 0 12px}h2{font-size:18px;margin:24px 0 12px}p{margin:7px 0;white-space:pre-wrap}
.note{font-size:12px;color:#667085;padding-bottom:16px;border-bottom:1px solid #e4e7ec;margin-bottom:22px}
.table-wrap{overflow-x:auto;margin:14px 0}table{border-collapse:collapse;width:100%;font-size:13px}td,th{border:1px solid #b8c4cc;padding:7px 9px;min-width:55px;vertical-align:top}
td p{margin:3px 0}img{max-width:100%;height:auto;display:block;margin:12px auto}ins{background:#eaf9f0;text-decoration:underline}del{background:#fff0ee;color:#9d3535}
.unavailable{color:#906119;font-size:12px}.sheet-label{background:#edf8f2;padding:8px 12px}td{white-space:pre-wrap}
@media(max-width:600px){main{margin:0;padding:20px 16px}}
"""


def xml(data):
    if b"<!DOCTYPE" in data.upper() or b"<!ENTITY" in data.upper():
        raise ValueError("XML entities are not supported")
    return ET.fromstring(data)


def safe_office_zip(data):
    archive = zipfile.ZipFile(io.BytesIO(data))
    if sum(item.file_size for item in archive.infolist()) > 400 * 1024 * 1024:
        raise ValueError("Office document exceeds preview size limit")
    return archive


def docx_preview(data):
    with safe_office_zip(data) as archive:
        rels = {}
        if "word/_rels/document.xml.rels" in archive.namelist():
            for rel in xml(archive.read("word/_rels/document.xml.rels")):
                if rel.get("TargetMode") != "External":
                    target = posixpath.normpath(posixpath.join("word", rel.get("Target", "")))
                    if target.startswith("word/media/"):
                        rels[rel.get("Id")] = target

        def render(node):
            tag = node.tag
            if tag == "{http://schemas.openxmlformats.org/markup-compatibility/2006}AlternateContent":
                choice = next((child for child in node if child.tag.endswith('}Choice')), None)
                fallback = next((child for child in node if child.tag.endswith('}Fallback')), None)
                selected = choice if choice is not None else fallback
                return render(selected) if selected is not None else ""
            if tag in (W + "pPr", W + "rPr", W + "tblPr", W + "tcPr", W + "sectPr", W + "instrText"):
                return ""
            if tag in (W + "t", W + "delText"):
                return html.escape(node.text or "")
            if tag == W + "tab":
                return "　"
            if tag in (W + "br", W + "cr"):
                return "<br>"
            if tag in (A + "blip", V + "imagedata"):
                target = rels.get(node.get(R + "embed") or node.get(R + "id"))
                if target and target in archive.namelist():
                    mime = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".gif": "image/gif", ".bmp": "image/bmp", ".webp": "image/webp"}.get(Path(target).suffix.lower())
                    if mime:
                        encoded = base64.b64encode(archive.read(target)).decode("ascii")
                        return f'<img loading="lazy" alt="原文档内嵌图片" src="data:{mime};base64,{encoded}">'
                return '<span class="unavailable">[此图片格式或外部图片无法预览，请查看原件]</span>'
            if tag == W + "tbl":
                rows = []
                # Preserve horizontal and vertical merged cells without trusting source HTML.
                active = {}
                for row in node.findall(W + "tr"):
                    cells = []
                    column = 0
                    next_active = {}
                    for cell in row.findall(W + "tc"):
                        span_el = cell.find(W + "tcPr/" + W + "gridSpan")
                        span = max(1, min(100, int(span_el.get(W + "val", "1")))) if span_el is not None else 1
                        merge = cell.find(W + "tcPr/" + W + "vMerge")
                        if merge is not None and merge.get(W + "val") != "restart" and column in active:
                            saved = active[column]
                            saved["rows"] += 1
                            if any(child.tag in (W + "t", W + "delText", A + "blip", V + "imagedata") for child in cell.iter()):
                                saved["content"] += "".join(render(child) for child in cell)
                            next_active[column] = saved
                        else:
                            saved = {"content": "".join(render(child) for child in cell), "span": span, "rows": 1}
                            cells.append(saved)
                            if merge is not None:
                                next_active[column] = saved
                        column += span
                    active = next_active
                    rows.append(cells)
                return '<div class="table-wrap"><table>' + "".join("<tr>" + "".join(f'<td colspan="{c["span"]}" rowspan="{c["rows"]}">{c["content"]}</td>' for c in row) + "</tr>" for row in rows) + "</table></div>"
            content = "".join(render(child) for child in node)
            if tag == W + "p":
                alignment = node.find(W + "pPr/" + W + "jc")
                value = alignment.get(W + "val") if alignment is not None else ""
                style = f' style="text-align:{value}"' if value in ("center", "right", "left") else ""
                return f"<p{style}>{content or '<br>'}</p>"
            if tag == W + "r":
                def enabled(name):
                    element = node.find(W + "rPr/" + W + name)
                    return element is not None and element.get(W + "val", "1") not in ("0", "false", "off")
                if enabled("b"):
                    content = f"<strong>{content}</strong>"
                if enabled("i"):
                    content = f"<em>{content}</em>"
            if tag in (W + "ins", W + "del"):
                wrapper = "ins" if tag == W + "ins" else "del"
                return f"<{wrapper}>{content}</{wrapper}>"
            return content

        body = xml(archive.read("word/document.xml")).find(W + "body")
        return render(body), "内容预览保留正文、表格、可读取图片和修订标记；不等同于 Word 分页效果，页眉页脚等请查看原件。"


def docx_model(data):
    """Extract safe text and raster-image blocks for the in-app Office workspace.

    DOCX does not contain reliable rendered page boundaries. Explicit page
    breaks are honoured; otherwise the document is exposed as continuous
    content and the original HTML preview remains available for visual QA.
    """
    with safe_office_zip(data) as archive:
        media_relationships = {}
        if "word/_rels/document.xml.rels" in archive.namelist():
            for rel in xml(archive.read("word/_rels/document.xml.rels")):
                if rel.get("TargetMode") != "External":
                    target = posixpath.normpath(posixpath.join("word", rel.get("Target", "")))
                    if target.startswith("word/media/"):
                        media_relationships[rel.get("Id")] = target
        styles = {}
        if "word/styles.xml" in archive.namelist():
            for style in xml(archive.read("word/styles.xml")).findall(W + "style"):
                style_id = style.get(W + "styleId")
                name = style.find(W + "name")
                if style_id:
                    styles[style_id] = name.get(W + "val", style_id) if name is not None else style_id

        body = xml(archive.read("word/document.xml")).find(W + "body")
        pages = []
        blocks = []
        outline = []
        page_number = 1
        block_number = 0

        def flush_page():
            nonlocal blocks, page_number
            if not blocks:
                return
            pages.append({
                "id": f"page-{page_number}",
                "label": "连续内容" if page_number == 1 and not pages else f"第 {page_number} 页",
                "blocks": blocks,
            })
            blocks = []
            page_number += 1

        def paragraph_text(paragraph):
            parts = []
            for node in paragraph.iter():
                if node.tag in (W + "t", W + "delText"):
                    parts.append(node.text or "")
                elif node.tag == W + "tab":
                    parts.append("\t")
                elif node.tag in (W + "br", W + "cr") and node.get(W + "type") != "page":
                    parts.append("\n")
            return "".join(parts).strip()

        def add_text_paragraph(paragraph, forced_text=None):
            nonlocal block_number
            text = forced_text if forced_text is not None else paragraph_text(paragraph)
            if not text:
                return
            block_number += 1
            block_id = f"block-{block_number}"
            props = paragraph.find(W + "pPr")
            style_id = None
            outline_level = None
            numbered = False
            if props is not None:
                style = props.find(W + "pStyle")
                style_id = style.get(W + "val") if style is not None else None
                outline_node = props.find(W + "outlineLvl")
                if outline_node is not None:
                    try:
                        outline_level = int(outline_node.get(W + "val", "0")) + 1
                    except ValueError:
                        outline_level = None
                numbered = props.find(W + "numPr") is not None
            style_name = styles.get(style_id, style_id or "")
            normalized_style = style_name.casefold()
            is_heading = outline_level is not None or "heading" in normalized_style or "标题" in style_name
            level = min(3, max(1, outline_level or next(
                (int(char) for char in style_name if char in "123"),
                1,
            ))) if is_heading else None
            block = {
                "id": block_id,
                "text": text,
                "type": "heading" if is_heading else "list-item" if numbered else "paragraph",
            }
            if level is not None:
                block["level"] = level
                outline.append({
                    "id": f"outline-{block_number}",
                    "blockId": block_id,
                    "label": text,
                    "level": level,
                    "pageId": f"page-{page_number}",
                })
            blocks.append(block)

        def paragraph_segments(paragraph):
            """Keep mixed text/image order; never duplicate AlternateContent fallbacks."""
            pending = []
            segments = []

            def flush():
                text = "".join(pending).strip()
                pending.clear()
                if text:
                    segments.append(("text", text))

            def walk(node):
                if node.tag == "{http://schemas.openxmlformats.org/markup-compatibility/2006}AlternateContent":
                    choice = next((child for child in node if child.tag.endswith("}Choice")), None)
                    fallback = next((child for child in node if child.tag.endswith("}Fallback")), None)
                    selected = choice if choice is not None else fallback
                    if selected is not None:
                        walk(selected)
                    return
                if node.tag in (W + "pPr", W + "rPr", W + "instrText"):
                    return
                if node.tag in (W + "t", W + "delText"):
                    pending.append(node.text or "")
                elif node.tag == W + "tab":
                    pending.append("\t")
                elif node.tag in (W + "br", W + "cr") and node.get(W + "type") != "page":
                    pending.append("\n")
                elif node.tag in (A + "blip", V + "imagedata"):
                    flush()
                    target = media_relationships.get(node.get(R + "embed") or node.get(R + "id"))
                    image = {"alt": "原文档内嵌图片"}
                    if target and target in archive.namelist():
                        mime = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".gif": "image/gif",
                                ".bmp": "image/bmp", ".webp": "image/webp"}.get(Path(target).suffix.lower())
                        if mime and archive.getinfo(target).file_size <= 20 * 1024 * 1024:
                            image["src"] = f"data:{mime};base64," + base64.b64encode(archive.read(target)).decode("ascii")
                    segments.append(("image", image))
                else:
                    for child in node:
                        walk(child)

            walk(paragraph)
            flush()
            return segments

        def add_paragraph(paragraph):
            nonlocal block_number
            for kind, value in paragraph_segments(paragraph):
                if kind == "text":
                    add_text_paragraph(paragraph, value)
                else:
                    block_number += 1
                    blocks.append({"id": f"block-{block_number}", "type": "image", "text": "", "image": value})

        if body is not None:
            for child in body:
                if child.tag == W + "p":
                    add_paragraph(child)
                    has_page_break = any(
                        node.tag == W + "lastRenderedPageBreak"
                        or (node.tag == W + "br" and node.get(W + "type") == "page")
                        for node in child.iter()
                    )
                    if has_page_break:
                        flush_page()
                elif child.tag == W + "tbl":
                    if any(node.tag in (A + "blip", V + "imagedata") for node in child.iter()):
                        # This fallback is continuous content, not an Office table
                        # renderer. Preserve cell reading order and every picture.
                        for paragraph in child.iter(W + "p"):
                            add_paragraph(paragraph)
                        continue
                    rows = []
                    for row in child.findall(W + "tr"):
                        cells = [paragraph_text(cell) for cell in row.findall(W + "tc")]
                        if any(cells):
                            rows.append(" | ".join(cells))
                    if rows:
                        add_text_paragraph(child, "\n".join(rows))
        flush_page()
        if not pages:
            pages = [{"id": "page-1", "label": "连续内容", "blocks": []}]
        return {"wordDocument": {"pages": pages, "outline": outline}}


def xlsx_preview(data):
    with safe_office_zip(data):
        pass
    values = load_workbook(io.BytesIO(data), data_only=True, keep_links=False)
    formulas = load_workbook(io.BytesIO(data), data_only=False, keep_links=False)
    parts = []
    try:
        for sheet in values:
            if sheet.max_row * sheet.max_column > 100000:
                raise ValueError("Worksheet exceeds preview cell limit")
            skipped = set()
            spans = {}
            for merged in sheet.merged_cells.ranges:
                spans[(merged.min_row, merged.min_col)] = (merged.max_row - merged.min_row + 1, merged.max_col - merged.min_col + 1)
                skipped.update((r, c) for r in range(merged.min_row, merged.max_row + 1) for c in range(merged.min_col, merged.max_col + 1) if (r, c) != (merged.min_row, merged.min_col))
            rows = []
            for row in sheet:
                cells = []
                for cell in row:
                    if (cell.row, cell.column) in skipped:
                        continue
                    value = cell.value
                    formula = formulas[sheet.title].cell(cell.row, cell.column)
                    if value is None and formula.data_type == "f":
                        value = f"{formula.value}（无缓存结果）"
                    elif isinstance(value, (int, float)) and not isinstance(value, bool):
                        fmt = cell.number_format
                        if "%" in fmt:
                            decimals = len(fmt.split(".")[-1].split("%")[0]) if "." in fmt else 0
                            value = f"{value:.{min(decimals, 8)}%}"
                        elif "." in fmt and "0" in fmt:
                            decimals = len(fmt.split(".")[-1].split(";")[0].rstrip('"'))
                            value = f"{value:,.{min(decimals, 8)}f}" if "," in fmt else f"{value:.{min(decimals, 8)}f}"
                    if hasattr(value, "isoformat"):
                        value = value.isoformat(sep=" ")
                    rowspan, colspan = spans.get((cell.row, cell.column), (1, 1))
                    content = html.escape(str(value)) if value is not None else ""
                    if cell.font.bold:
                        content = f"<strong>{content}</strong>"
                    cells.append(f'<td rowspan="{rowspan}" colspan="{colspan}">{content}</td>')
                rows.append("<tr>" + "".join(cells) + "</tr>")
            parts.append(f'<h2 class="sheet-label">{html.escape(sheet.title)}</h2><div class="table-wrap"><table>' + "".join(rows) + "</table></div>")
    finally:
        values.close()
        formulas.close()
    return "".join(parts), "只读展示工作表及原文件缓存结果；未重算公式，无缓存的公式按原式显示。格式与打印效果以 Excel 原件为准。"


def xlsx_model(data):
    """Extract workbook sheets and formulas without executing workbook code."""
    with safe_office_zip(data):
        pass
    values = load_workbook(io.BytesIO(data), data_only=True, keep_links=False)
    formulas = load_workbook(io.BytesIO(data), data_only=False, keep_links=False)
    sheets = []
    try:
        for index, sheet in enumerate(values):
            if sheet.max_row * sheet.max_column > 100000:
                raise ValueError("Worksheet exceeds preview cell limit")
            rows = []
            for row_index in range(1, sheet.max_row + 1):
                row = []
                for column_index in range(1, sheet.max_column + 1):
                    value = sheet.cell(row_index, column_index).value
                    formula_value = formulas[sheet.title].cell(row_index, column_index).value
                    if hasattr(value, "isoformat"):
                        value = value.isoformat(sep=" ")
                    if isinstance(formula_value, str) and formula_value.startswith("="):
                        row.append({"value": value, "formula": formula_value})
                    elif value is None or isinstance(value, (str, int, float, bool)):
                        row.append(value)
                    else:
                        row.append(str(value))
                while row and row[-1] is None:
                    row.pop()
                rows.append(row)
            while rows and not rows[-1]:
                rows.pop()
            sheets.append({"id": f"sheet-{index + 1}", "name": sheet.title, "rows": rows})
    finally:
        values.close()
        formulas.close()
    return {"workbook": {"sheets": sheets}}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("archive", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    package = args.archive.read_bytes()
    digest = hashlib.sha256(package).hexdigest()
    directory = args.output / digest[:12]
    directory.mkdir(parents=True, exist_ok=True)
    result = {"projectId": "207", "taskId": "3499", "source": args.archive.name, "sha256": digest, "directory": digest[:12], "files": []}
    with zipfile.ZipFile(io.BytesIO(package)) as archive:
        source = json.loads(archive.read("manifest.json"))
        if source.get("project_id") != 207 or source.get("task_id") != 3499:
            raise ValueError("Package project/task does not match the approved local preview")
        entries = [e for e in archive.infolist() if e.filename.split("/")[0] in GROUPS and Path(e.filename).suffix.lower() in (".docx", ".xlsx")]
        if len(entries) != 9 or sum(e.file_size for e in entries) > 200 * 1024 * 1024:
            raise ValueError("Unexpected package contents")
        def entry_order(entry):
            name = entry.filename.split('/')[-1]
            ordinal = next((i for i, token in enumerate(['（一）', '（二）', '（三）', '（四）']) if name.startswith(token)), 99)
            return list(GROUPS).index(entry.filename.split('/')[0]), ordinal, name
        for index, entry in enumerate(sorted(entries, key=entry_order)):
            data = archive.read(entry)
            name = entry.filename.split("/")[-1]
            extension = Path(name).suffix.lower()
            file_id = f"file-{index + 1}"
            (directory / f"{file_id}{extension}").write_bytes(data)
            body, note = docx_preview(data) if extension == ".docx" else xlsx_preview(data)
            model = docx_model(data) if extension == ".docx" else xlsx_model(data)
            document = '<!doctype html><html lang="zh-CN"><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>' + html.escape(name) + '</title><style>' + STYLE + '</style><main><h1>' + html.escape(name) + '</h1><p class="note">' + html.escape(note) + '</p>' + body + '</main></html>'
            (directory / f"{file_id}.html").write_text(document, encoding="utf-8")
            (directory / f"{file_id}.json").write_text(json.dumps(model, ensure_ascii=False), encoding="utf-8")
            result["files"].append({"id": file_id, "name": name, "category": GROUPS[entry.filename.split('/')[0]], "size": len(data), "extension": extension, "sha256": hashlib.sha256(data).hexdigest()})
    (args.output / "manifest.json").write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"projectId": result["projectId"], "taskId": result["taskId"], "files": len(result["files"]), "output": str(args.output)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
